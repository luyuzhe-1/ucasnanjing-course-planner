import json
import re
from collections import Counter
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data-source" / "2026-2027学年计划开设课程总表.xlsx"
SPRING_REFERENCE = ROOT / "data-source" / "2025-2026春季学期课表.pdf"
OUTPUT = ROOT / "app" / "data" / "courses.json"
DAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def text(value):
    return str(value).strip() if value is not None else ""


def normalize_code(value):
    value = re.sub(r"^280223", "", text(value))
    value = value.split("-", 1)[0]
    return re.sub(r"[（(].*?[）)]", "", value).strip()


def normalize_name(value):
    value = text(value).replace("（", "(").replace("）", ")")
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"\((线上课[^)]*|线)\)", "", value)
    return value


def semester_list(term):
    return [semester for semester in ("秋", "春") if semester in term]


def expand_weeks(spec):
    weeks = set()
    for part in spec.replace("第", "").replace("周", "").replace("、", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start, end = (int(value) for value in part.split("-", 1))
            weeks.update(range(start, end + 1))
        elif part.isdigit():
            weeks.add(int(part))
    return sorted(weeks)


def parse_schedule_lines(cell):
    parsed = []
    pending_name = ""
    for raw_line in (cell or "").splitlines():
        line = raw_line.strip().replace("，", ",").replace("、", ",")
        if not line:
            continue
        if pending_name and re.match(r"^第?\d", line):
            line = f"{pending_name},{line}"
            pending_name = ""
        match = re.match(r"^(.*?),第?([0-9,-]+周)(?:,(.+))?$", line)
        if match:
            name, week_spec, room = match.groups()
            parsed.append((name.strip(), week_spec, (room or "").strip()))
        elif "," not in line:
            pending_name = line
    return parsed


def load_spring_reference_schedule():
    with pdfplumber.open(SPRING_REFERENCE) as pdf:
        table = pdf.pages[0].extract_table({
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "intersection_tolerance": 3,
            "snap_tolerance": 3,
        })
    row_periods = {1: 1, 2: 2, 3: 3, 4: 4, 6: 5, 7: 6, 8: 7, 9: 8,
                   10: 9, 12: 10, 13: 11, 14: 12, 15: 13}
    occurrences = {}
    for day_index, day in enumerate(DAY_NAMES, start=2):
        active_entries = []
        for row_index, period in row_periods.items():
            cell = table[row_index][day_index]
            if cell is None:
                entries = active_entries
            else:
                entries = parse_schedule_lines(cell)
                active_entries = entries
            for name, week_spec, room in entries:
                key = (normalize_name(name), name, day, week_spec, room)
                occurrences.setdefault(key, set()).add(period)
    schedule = {}
    for (normalized, name, day, week_spec, room), periods in occurrences.items():
        schedule.setdefault(normalized, []).append({
            "name": name,
            "day": day,
            "dayIndex": DAY_NAMES.index(day) + 1,
            "periods": sorted(periods),
            "weekSpec": week_spec,
            "weeks": expand_weeks(week_spec),
            "room": room,
            "semester": "春",
            "reference": True,
            "source": "2025–2026春季学期课表",
        })
    return schedule


def load_plan():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [text(value) for value in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    rows = []
    current_school = ""
    for values in sheet.iter_rows(min_row=3, values_only=True):
        record = dict(zip(headers, values))
        if record.get("开课单位"):
            current_school = text(record["开课单位"])
        record = {key: text(value) for key, value in record.items()}
        record["开课单位"] = current_school
        if record.get("课程编码") and record.get("课程中文名称"):
            rows.append(record)
    return rows


def find_plan(course, plan_rows):
    code = normalize_code(course.get("code") or course.get("id"))
    by_code = [row for row in plan_rows if normalize_code(row["课程编码"]) == code]
    if by_code:
        return by_code[0]
    name = normalize_name(course.get("name"))
    by_name = [row for row in plan_rows if normalize_name(row["课程中文名称"]) == name]
    return by_name[0] if by_name else None


def main():
    plan_rows = load_plan()
    spring_reference = load_spring_reference_schedule()
    existing = json.loads(OUTPUT.read_text(encoding="utf-8"))
    courses = []
    represented_codes = set()

    for course in existing["courses"]:
        # Rebuilding is idempotent: retain only the current-year autumn schedule.
        course["meetings"] = [meeting for meeting in course.get("meetings", []) if not meeting.get("reference")]
        for meeting in course["meetings"]:
            meeting["semester"] = "秋"
            meeting["reference"] = False
            meeting["source"] = "2026–2027秋季学期正式课表"
        row = find_plan(course, plan_rows)
        if row:
            represented_codes.add(normalize_code(row["课程编码"]))
            course.update({
                "school": row["开课单位"],
                "discipline": row["课程所属学科"],
                "type": row["课程属性"],
                "level": row["培养层次"],
                "hours": row["课程学时"],
                "credits": row["课程学分"],
                "plannedTerm": row["开课学期"],
                "note": row["备注"],
            })
        course["semesters"] = semester_list(course.get("plannedTerm") or course.get("term", "秋"))
        if "春" in course["semesters"]:
            course["meetings"].extend(spring_reference.get(normalize_name(course["name"]), []))
        course["scheduleStatus"] = "已排课" if any(not meeting["reference"] for meeting in course["meetings"]) else ("参考排课" if course["meetings"] else "待排课")
        courses.append(course)

    for row in plan_rows:
        code = normalize_code(row["课程编码"])
        if code in represented_codes:
            continue
        reference_meetings = spring_reference.get(normalize_name(row["课程中文名称"]), []) if "春" in semester_list(row["开课学期"]) else []
        courses.append({
            "id": row["课程编码"],
            "name": row["课程中文名称"],
            "englishName": "",
            "code": row["课程编码"],
            "school": row["开课单位"],
            "discipline": row["课程所属学科"],
            "type": row["课程属性"],
            "level": row["培养层次"],
            "hours": row["课程学时"],
            "credits": row["课程学分"],
            "term": row["开课学期"],
            "plannedTerm": row["开课学期"],
            "semesters": semester_list(row["开课学期"]),
            "note": row["备注"],
            "limit": "",
            "enrolled": "",
            "teachingMethod": "",
            "examMethod": "",
            "chiefProfessor": "",
            "teacher": "",
            "scheduleStatus": "参考排课" if reference_meetings else "待排课",
            "meetings": reference_meetings,
        })

    term_counts = Counter()
    for row in plan_rows:
        for semester in semester_list(row["开课学期"]):
            term_counts[semester] += 1

    existing_meta = existing["meta"]
    output = {
        "meta": {
            "title": "2026–2027学年",
            "source": "2026-2027学年计划开设课程总表",
            "courseCount": len(plan_rows),
            "catalogEntryCount": len(courses),
            "semesterCounts": dict(term_counts),
            "scheduledCourseCount": sum(1 for course in courses if course["meetings"]),
            "springReferenceCourseCount": sum(1 for course in courses if any(meeting.get("reference") for meeting in course["meetings"])),
            "springReferenceMeetingCount": sum(1 for course in courses for meeting in course["meetings"] if meeting.get("reference")),
            "meetingCount": sum(len(course["meetings"]) for course in courses),
            "periodTimes": existing_meta["periodTimes"],
        },
        "courses": courses,
    }
    OUTPUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(output["meta"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
